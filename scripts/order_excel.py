#! /usr/bin/env python
# -*- coding:utf-8 -*-

import pymysql
import openpyxl
import sys
import time, datetime

def get_newadd_order(date1, date2):
	conn = pymysql.connect(host='10.23.210.26', port=3306, user='user', passwd='app123',
					   		db='oss', charset='utf8')

	cursor = conn.cursor()
	
	# date1 = sys.argv[1]
	# date2 = sys.argv[2]
	
	sql = "SELECT oof.onumber, oof.usertoken, oof.userid, oof.pid, oof.pcode, oof.pname, oof.price, " \
		  "CASE WHEN oof.`status` = '0' THEN '有效' END, CASE WHEN oof.paytype = '1' THEN '微信' END, " \
		  "oof.ordertime, oof.canceltime, oof.begintime, oof.endtime, " \
		  "CASE WHEN oof.pisgoon = '1' THEN '否' WHEN oof.pisgoon = '0' THEN '是' END, " \
		  "oof.contentid, sm.content_name, oof.tariff_cycle FROM oss_order_form oof " \
		  "LEFT JOIN " \
		  "oss_servicemapping sm ON sm.content_id = oof.contentid " \
		  "WHERE " \
		  "oof.ordertime > '" + date1 + "' AND oof.ordertime < '" + date2 + "'"
          # + "' ORDER BY ordertime DESC"
	
	cursor.execute(sql)
 
	# 获取剩余结果的第一行数据
	# row_1 = cursor.fetchone()

	# 获取剩余结果前n行数据
	# row_2 = cursor.fetchmany(3)
	 
	# 获取剩余结果所有数据
	rows = cursor.fetchall()
	
	cursor.close()
	conn.close()

	return rows

def newadd_order():

    day = datetime.date.today().strftime('%Y-%m-%d')

    day_1 = datetime.date.today().replace(day=1).strftime('%Y-%m-%d')

    newadds = get_newadd_order(day_1, day)

    ex = openpyxl.load_workbook(r'D:\test.xlsx')

    s1 = ex.get_sheet_by_name('新增订单')

    ex.remove(row=2)
    for newadd in newadds:
        s1.append(newadd)

    ex.save(r'D:\test.xlsx')

    print('sheet1: done')


###################################################

def get_ordering():
    conn = pymysql.connect(host='10.23.210.26', port=3306, user='user', passwd='app123',
                           db='oss', charset='utf8')

    cursor = conn.cursor()

    day = datetime.date.today().strftime('%Y-%m-%d')

    # yesday = day - datetime.timedelta(days=1).strftime('%Y-%m-%d')

    sql = "SELECT oof.onumber, oof.usertoken, oof.userid, oof.pid, oof.pcode, oof.pname, oof.price, " \
          "CASE WHEN " \
          "oof.`status` = '0' THEN '有效' END, " \
          "CASE WHEN oof.paytype = '1' THEN '微信' END, " \
          "oof.ordertime, oof.canceltime, oof.begintime, oof.endtime, " \
          "CASE WHEN oof.pisgoon = '0' THEN '是' END, " \
          "oof.contentid, sm.content_name, oof.tariff_cycle " \
          "FROM oss_order_form oof " \
          "LEFT JOIN oss_servicemapping sm ON sm.content_id = oof.contentid " \
          "WHERE oof.endtime = '' " \
          "AND oof.pisgoon = 0 " \
          "AND oof.ordertime < '" + day + "'" + \
          "GROUP BY onumber ORDER BY ordertime"

    cursor.execute(sql)

    rows = cursor.fetchall()

    cursor.close()
    conn.close()

    return rows

def ordering():

    orders = get_ordering()

    ex = openpyxl.load_workbook(r'D:\test.xlsx')

    s2 = ex.get_sheet_by_name('续订订单')

    for order in orders:
        s2.append(order)

    ex.save(r'D:\test.xlsx')

    print('sheet2 done')

def clear_sheet():

    ex = openpyxl.load_workbook(r'D:\test.xlsx')

    s1 = ex['新增订单']

    s2 = ex['续订订单']

    s1.delete_rows(2, 65530)

    s2.delete_rows(2, 65530)

    ex.save(r'D:\test.xlsx')

    print("clear done")

clear_sheet()

#newadd_order()

#ordering()

